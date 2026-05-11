#!/usr/bin/env python3
"""
Timesheet daily hour registration tool.

Reads git commits for a given date, extracts DevOps task IDs from commit
messages, and appends rows to the Composables Timesheet CSV.

If no commits are found for the day, a single N/A row is registered.
On first run, prompts for the user name and saves it to .timesheet-config.json.

Usage:
    python timesheet.py                  # register today
    python timesheet.py --date 20260508  # register specific date (YYYYMMDD)
    python timesheet.py --config         # reconfigure user/settings
"""

import argparse
import ctypes
import ctypes.wintypes
import csv
import json
import os
import re
import smtplib
import subprocess
import sys
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import shutil
import tempfile
import time

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent.resolve()
CONFIG_FILE = SCRIPT_DIR / ".timesheet-config.json"
EMAIL_CONFIG_FILE = SCRIPT_DIR / ".email-config.json"
XLSX_FILE = SCRIPT_DIR / "Composables Timesheet.xlsx"
XLSX_LOCAL = SCRIPT_DIR / "Composables Timesheet - Leonardo.xlsx"
XLSX_SHEET = "Timesheet"

# ---------------------------------------------------------------------------
# Portuguese weekday names
# ---------------------------------------------------------------------------
WEEKDAY_PT = [
    "Segunda-feira",
    "Terça-feira",
    "Quarta-feira",
    "Quinta-feira",
    "Sexta-feira",
    "Sábado",
    "Domingo",
]

# Regex that matches 5-digit numbers in the DevOps ID range seen in the sheet
DEVOPS_ID_RE = re.compile(r"\b(2\d{4})\b")

# Regex for "Related work items: #NNNN" in commit body (Azure DevOps format)
RELATED_WORK_RE = re.compile(r"Related work items?:\s*(.+)", re.IGNORECASE)
HASH_ID_RE = re.compile(r"#(\d{4,6})")

# Optional explicit hours tag in commit message: [horas: 3] or [horas: 1.5]
HORAS_TAG_RE = re.compile(r"\[horas:\s*([\d.,]+)\]", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------
def load_config() -> dict:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_config(cfg: dict) -> None:
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    print(f"[timesheet] Configuração guardada em {CONFIG_FILE}")


def prompt_config(existing: dict) -> dict:
    """Interactively ask user to fill in configuration."""
    print("\n=== Configuração do Timesheet ===")

    user = input(
        f"Nome do utilizador [{existing.get('user', '')}]: "
    ).strip()
    if not user:
        user = existing.get("user", "")
    if not user:
        print("Erro: o nome do utilizador é obrigatório.")
        sys.exit(1)

    project = input(
        f"Projeto por omissão [{existing.get('project', 'SI Pessoas')}]: "
    ).strip() or existing.get("project", "SI Pessoas")

    task = input(
        f"Tipo de tarefa por omissão [{existing.get('default_task', 'Dev')}]: "
    ).strip() or existing.get("default_task", "Dev")

    hours_input = input(
        f"Horas por dia por omissão [{existing.get('default_hours', 8.0)}]: "
    ).strip()
    try:
        hours = float(hours_input) if hours_input else float(existing.get("default_hours", 8.0))
    except ValueError:
        hours = 8.0

    existing_repos = existing.get("git_repos", existing.get("git_repo_path", "."))
    if isinstance(existing_repos, list):
        existing_repos_display = ", ".join(existing_repos)
    else:
        existing_repos_display = existing_repos
    print(f"Repositórios git actuais: {existing_repos_display}")
    print("Introduz os caminhos dos repositórios, um por linha.")
    print("Deixa uma linha em branco para terminar (mantém os actuais se não introduzires nada):")
    new_repos = []
    while True:
        r = input("  Repositório: ").strip()
        if not r:
            break
        new_repos.append(r)
    if not new_repos:
        # keep existing
        if isinstance(existing_repos, list):
            git_repos = existing_repos
        else:
            git_repos = [existing_repos] if existing_repos else ["."]
    else:
        git_repos = new_repos

    git_author = input(
        f"Email/nome do autor git (deixar vazio = todos) [{existing.get('git_author', '')}]: "
    ).strip() or existing.get("git_author", "")

    print("\n--- Configuração de Email (deixa em branco para desativar envio) ---")
    email_to = input(
        f"Destinatário do resumo (email) [{existing.get('email_to', '')}]: "
    ).strip() or existing.get("email_to", "")

    smtp_host = ""
    smtp_port = 587
    smtp_user = ""
    smtp_password = ""
    if email_to:
        smtp_host = input(
            f"Servidor SMTP [{existing.get('smtp_host', 'smtp.office365.com')}]: "
        ).strip() or existing.get("smtp_host", "smtp.office365.com")
        smtp_port_input = input(
            f"Porta SMTP [{existing.get('smtp_port', 587)}]: "
        ).strip()
        try:
            smtp_port = int(smtp_port_input) if smtp_port_input else int(existing.get("smtp_port", 587))
        except ValueError:
            smtp_port = 587
        smtp_user = input(
            f"Email remetente (From) [{existing.get('smtp_user', '')}]: "
        ).strip() or existing.get("smtp_user", "")
        smtp_password = input(
            f"Palavra-passe SMTP (guardada em texto simples) [{('*' * 8) if existing.get('smtp_password') else ''}]: "
        ).strip() or existing.get("smtp_password", "")

    return {
        "user": user,
        "project": project,
        "default_task": task,
        "default_hours": hours,
        "git_repos": git_repos,
        "git_author": git_author,
        "email_to": email_to,
        "smtp_host": smtp_host,
        "smtp_port": smtp_port,
        "smtp_user": smtp_user,
        "smtp_password": smtp_password,
    }


def ensure_config(force: bool = False) -> dict:
    cfg = load_config()
    if force or not cfg.get("user"):
        cfg = prompt_config(cfg)
        save_config(cfg)
    return cfg


# ---------------------------------------------------------------------------
# Git helpers
# ---------------------------------------------------------------------------
def get_commits_from_repo(repo_path: str, target_date: date, author: str) -> list[dict]:
    """
    Return commits from a single repo on target_date.
    Each item: {"hash": str, "subject": str, "datetime": str, "repo": str}
    """
    after = f"{target_date.strftime('%Y-%m-%d')} 00:00:00"
    before = f"{target_date.strftime('%Y-%m-%d')} 23:59:59"

    cmd = [
        "git", "-C", repo_path,
        "log",
        f"--after={after}",
        f"--before={before}",
        "--format=%x1e%ai%x00%H%x00%s%x00%b",
        "--no-merges",
    ]
    if author:
        cmd += [f"--author={author}"]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except FileNotFoundError:
        print("[timesheet] Erro: comando 'git' não encontrado. Instala o Git e tenta novamente.")
        sys.exit(1)

    if result.returncode != 0:
        return []

    commits = []
    for record in result.stdout.split("\x1e"):
        record = record.strip()
        if not record:
            continue
        parts = record.split("\x00", 3)
        if len(parts) < 3:
            continue
        commits.append({
            "datetime": parts[0].strip(),
            "hash": parts[1].strip(),
            "subject": parts[2].strip(),
            "body": parts[3].strip() if len(parts) > 3 else "",
            "repo": repo_path,
        })
    return commits


def get_commits_for_date(repo_paths: list[str], target_date: date, author: str) -> list[dict]:
    """
    Aggregate commits from all configured repositories for target_date.
    Deduplicates by commit hash (same commit can appear if repos share history).
    """
    seen_hashes: set[str] = set()
    all_commits: list[dict] = []
    for repo in repo_paths:
        repo = repo.strip()
        if not repo:
            continue
        commits = get_commits_from_repo(repo, target_date, author)
        for c in commits:
            if c["hash"] not in seen_hashes:
                seen_hashes.add(c["hash"])
                all_commits.append(c)
    return all_commits


def get_commits_for_date_extended(repo_paths: list[str], target_date: date, author: str) -> list[dict]:
    """
    Like get_commits_for_date, but when target_date is a Monday also includes
    commits from the preceding Saturday and Sunday (registered under Monday).
    """
    commits = get_commits_for_date(repo_paths, target_date, author)
    if target_date.weekday() == 0:  # Monday
        seen = {c["hash"] for c in commits}
        weekend_count = 0
        for delta in (2, 1):  # Saturday (-2 days), then Sunday (-1 day)
            weekend_date = target_date - timedelta(days=delta)
            for c in get_commits_for_date(repo_paths, weekend_date, author):
                if c["hash"] not in seen:
                    commits.append(c)
                    seen.add(c["hash"])
                    weekend_count += 1
        if weekend_count:
            print(f"[timesheet] Segunda-feira: {weekend_count} commit(s) de fim de semana incluídos em {target_date.strftime('%d/%m/%Y')}.")
    return commits


# ---------------------------------------------------------------------------
# DevOps ID extraction
# ---------------------------------------------------------------------------
def extract_devops_ids(subject: str, body: str = "") -> list[str]:
    # 1. Look for "Related work items: #NNNN" in commit body (Azure DevOps PR merges)
    for line in body.splitlines():
        m = RELATED_WORK_RE.match(line.strip())
        if m:
            ids = HASH_ID_RE.findall(m.group(1))
            if ids:
                return ids
    # 2. Fall back to 5-digit numbers starting with 2 in subject
    return DEVOPS_ID_RE.findall(subject)


# ---------------------------------------------------------------------------
# Excel (openpyxl) helpers
# ---------------------------------------------------------------------------
_RETRIES = 8
_RETRY_DELAY = 5.0  # seconds

# Windows API constants for shared file access
_GENERIC_READ = 0x80000000
_FILE_SHARE_ALL = 0x1 | 0x2 | 0x4   # READ | WRITE | DELETE
_OPEN_EXISTING = 3
_FILE_ATTR_NORMAL = 0x80
_INVALID_HANDLE = ctypes.wintypes.HANDLE(-1).value


def _read_file_shared(path: Path) -> bytes:
    """
    Read file bytes using Win32 CreateFile with full share flags.
    Works even when Excel/OneDrive holds an exclusive lock.
    """
    kernel32 = ctypes.windll.kernel32
    handle = kernel32.CreateFileW(
        str(path),
        _GENERIC_READ,
        _FILE_SHARE_ALL,
        None,
        _OPEN_EXISTING,
        _FILE_ATTR_NORMAL,
        None,
    )
    if handle == _INVALID_HANDLE:
        err = ctypes.get_last_error()
        raise PermissionError(
            f"Não foi possível abrir '{path}' (Win32 erro {err}).\n"
            "Verifica se o ficheiro existe e não está exclusivamente bloqueado."
        )
    try:
        size = kernel32.GetFileSize(handle, None)
        buf = ctypes.create_string_buffer(size)
        read = ctypes.c_ulong(0)
        kernel32.ReadFile(handle, buf, size, ctypes.byref(read), None)
        return buf.raw[: read.value]
    finally:
        kernel32.CloseHandle(handle)


def _openpyxl_load(xlsx_path: Path, read_only: bool = False):
    """
    Load workbook. First tries the normal path; if locked (PermissionError),
    reads raw bytes via Win32 shared-access API (bypasses Excel/OneDrive lock)
    and loads from a BytesIO / temp file.
    Returns (workbook, temp_path_or_None).
    """
    try:
        return openpyxl.load_workbook(xlsx_path, read_only=read_only), None
    except PermissionError:
        pass  # fall through to Win32 path

    data = _read_file_shared(xlsx_path)
    if read_only:
        import io as _io
        return openpyxl.load_workbook(_io.BytesIO(data), read_only=True), None
    # For write mode we need a real file path so openpyxl can save back
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    tmp.write_bytes(data)
    return openpyxl.load_workbook(tmp, read_only=False), tmp


def _openpyxl_save(wb, xlsx_path: Path, tmp_src: "Path | None" = None) -> None:
    """
    Save workbook back to xlsx_path with retry loop.
    If a temp copy was used to load, it is cleaned up after saving.
    If saving always fails, the modified workbook is kept in a temp file
    and the user is informed.
    """
    for attempt in range(1, _RETRIES + 1):
        try:
            wb.save(xlsx_path)
            if tmp_src and tmp_src.exists():
                tmp_src.unlink()
            print("[timesheet] Ficheiro guardado via openpyxl.")
            return
        except PermissionError:
            if attempt == _RETRIES:
                # Keep modified file in temp so data is not lost
                fallback = Path(tempfile.mktemp(suffix=".xlsx"))
                wb.save(fallback)
                print(f"[timesheet] AVISO: não foi possível guardar em:\n"
                      f"            {xlsx_path}")
                print(f"[timesheet] Cópia modificada guardada em:\n"
                      f"            {fallback}")
                print("[timesheet] Fecha o Excel, substitui o ficheiro original "
                      "pela cópia acima e volta a correr.")
                return
            print(f"[timesheet] Ficheiro bloqueado (tentativa {attempt}/{_RETRIES}): "
                  f"fecha o Excel... a aguardar {int(_RETRY_DELAY)}s")
            time.sleep(_RETRY_DELAY)


# --- Internal worksheet-level operations (work on an already-open ws) --------

def _ws_read_existing_dates(ws) -> set[str]:
    """Read 'YYYYMMDD|Pessoa' pairs from an open worksheet."""
    existing: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        col_a, col_b = row[0], row[1]
        if col_a is not None and col_b is not None:
            a = str(int(col_a)) if isinstance(col_a, float) else str(col_a).strip()
            existing.add(f"{a}|{str(col_b).strip()}")
    return existing


def _ws_get_last_registered_date(ws, user: str) -> "date | None":
    """Return last registered date for user from an open worksheet."""
    last: "date | None" = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        val_a, val_b = row[0], row[1]
        if val_b is None or str(val_b).strip() != user:
            continue
        if val_a is None:
            continue
        date_str = str(int(val_a)) if isinstance(val_a, float) else str(val_a).strip()
        if len(date_str) == 8 and date_str.isdigit():
            try:
                last = datetime.strptime(date_str, "%Y%m%d").date()
            except ValueError:
                pass
    return last


def _ws_delete_rows_for_date(ws, date_str: str, user: str) -> int:
    """Delete rows matching date_str+user. Returns count deleted."""
    to_delete = []
    for row in ws.iter_rows(min_row=2, max_col=2):
        val_a = row[0].value
        val_b = row[1].value
        a = str(int(val_a)) if isinstance(val_a, float) else str(val_a or "").strip()
        b = str(val_b or "").strip()
        if a == date_str and b == user:
            to_delete.append(row[0].row)
    for i in reversed(to_delete):
        ws.delete_rows(i)
    return len(to_delete)


def _ws_last_data_row(ws) -> int:
    """Return the index of the last row that has a non-empty value in column A."""
    last = 1
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip() != "":
            last = cell.row
    return last


def _ws_append_rows(ws, rows: list[list]) -> None:
    """Append typed rows immediately after the last row with data in column A."""
    next_row = _ws_last_data_row(ws) + 1
    for row in rows:
        for col_idx, val in enumerate(row, start=1):
            if col_idx == 1:  # date YYYYMMDD -> int
                try: val = int(val)
                except (ValueError, TypeError): pass
            elif col_idx == 4:  # DevOps ID -> int if numeric
                try: val = int(val)
                except (ValueError, TypeError): pass
            elif col_idx == 6:  # Hours -> float
                try: val = float(val)
                except (ValueError, TypeError): pass
            ws.cell(row=next_row, column=col_idx).value = val if val != "" else None
        next_row += 1


# --- Public functions (open+close Excel per call — used by tests) ------------

def read_existing_dates(xlsx_path: Path) -> set[str]:
    """Return set of 'YYYYMMDD|Pessoa' already in the xlsx."""
    if not xlsx_path.exists():
        return set()
    wb, tmp = _openpyxl_load(xlsx_path, read_only=True)
    ws = wb[XLSX_SHEET]
    result = _ws_read_existing_dates(ws)
    wb.close()
    if tmp and tmp.exists():
        tmp.unlink()
    return result


def replace_rows_for_date(xlsx_path: Path, date_str: str, user: str, new_rows: list[list]) -> int:
    """Delete existing rows for date+user and append new rows. Returns count deleted."""
    wb, tmp = _openpyxl_load(xlsx_path)
    ws = wb[XLSX_SHEET]
    deleted = _ws_delete_rows_for_date(ws, date_str, user)
    _ws_append_rows(ws, new_rows)
    _openpyxl_save(wb, xlsx_path, tmp)
    return deleted


def append_rows(xlsx_path: Path, rows: list[list]) -> None:
    """Append rows to the xlsx."""
    wb, tmp = _openpyxl_load(xlsx_path)
    ws = wb[XLSX_SHEET]
    _ws_append_rows(ws, rows)
    _openpyxl_save(wb, xlsx_path, tmp)


def _excel_append_rows(ws, rows: list[list]) -> None:
    """Alias kept for backwards compat."""
    _ws_append_rows(ws, rows)



def _commit_summary(commits: list[dict]) -> str:
    if not commits:
        return ""
    subjects = list(dict.fromkeys(c["subject"] for c in commits))
    summary = ", ".join(s[:40] for s in subjects[:4])
    if len(subjects) > 4:
        summary += f" (+{len(subjects)-4} mais)"
    return summary


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------
def build_rows(
    target_date: date,
    user: str,
    project: str,
    task: str,
    total_hours: float,
    commits: list[dict],
) -> list[list]:
    date_yyyymmdd = target_date.strftime("%Y%m%d")
    date_ddmmyyyy = target_date.strftime("%d-%m-%Y")
    weekday = WEEKDAY_PT[target_date.weekday()]

    if not commits:
        # No commits — register row with configured hours/task
        return [[
            date_yyyymmdd, user, project, "N/A", task, total_hours,
            "Sem commits", "", date_ddmmyyyy, weekday,
        ]]

    # Collect unique DevOps IDs across all commits
    task_map: dict[str, list[str]] = {}  # devops_id → [commit subjects]
    no_id_subjects: list[str] = []

    for commit in commits:
        ids = extract_devops_ids(commit["subject"], commit.get("body", ""))
        if ids:
            for did in ids:
                task_map.setdefault(did, []).append(commit["subject"])
        else:
            no_id_subjects.append(commit["subject"])

    # If no DevOps IDs found anywhere, group under "N/A"
    if not task_map:
        task_map["N/A"] = no_id_subjects
    elif no_id_subjects:
        # Attach orphan commits to the first found ID
        first_id = next(iter(task_map))
        task_map[first_id].extend(no_id_subjects)

    # Extract explicit [horas: X] tags — first match per task wins
    task_explicit_hours: dict[str, float] = {}
    for devops_id, subjects in task_map.items():
        for subject in subjects:
            m = HORAS_TAG_RE.search(subject)
            if m:
                try:
                    task_explicit_hours[devops_id] = float(m.group(1).replace(",", "."))
                    break
                except ValueError:
                    pass

    # Always cap total at 8h per day
    total_hours = min(total_hours, 8.0)
    explicit_total = sum(task_explicit_hours.values())
    tasks_without_tag = [tid for tid in task_map if tid not in task_explicit_hours]
    remaining = max(0.0, total_hours - explicit_total)

    # Distribute remaining hours as integers: base per task + 1 extra to first tasks
    task_implicit_hours: dict[str, float] = {}
    if tasks_without_tag:
        remaining_int = int(remaining)
        base = remaining_int // len(tasks_without_tag)
        extra = remaining_int % len(tasks_without_tag)
        for i, tid in enumerate(tasks_without_tag):
            task_implicit_hours[tid] = base + (1 if i < extra else 0)

    # Build rows
    rows = []
    task_ids = list(task_map.keys())
    for devops_id, subjects in task_map.items():
        comment = "; ".join(dict.fromkeys(subjects))  # deduplicate, preserve order
        comment = comment[:255]
        if devops_id in task_explicit_hours:
            h = task_explicit_hours[devops_id]
        else:
            h = task_implicit_hours.get(devops_id, 0)
        rows.append([
            date_yyyymmdd, user, project, devops_id, task,
            h, comment, "", date_ddmmyyyy, weekday,
        ])

    # Guarantee daily total = 8h: add any shortfall to the first row WITHOUT explicit hours
    current_total = sum(
        float(r[5]) for r in rows
        if r[5] not in (None, "", "N/A")
    )
    shortfall = round(total_hours - current_total, 2)
    if shortfall > 0:
        for row in rows:
            if row[3] not in task_explicit_hours:
                row[5] = round(float(row[5]) + shortfall, 2)
                break

    return rows


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------
def _build_monthly_excel(shared_xlsx: Path, user: str, year: int, month: int) -> bytes:
    """
    Return bytes of an xlsx with only the current month's rows for `user`,
    using the same columns as the shared file.
    """
    wb_src, tmp = _openpyxl_load(shared_xlsx, read_only=True)
    ws_src = wb_src[XLSX_SHEET]
    header = [cell.value for cell in next(ws_src.iter_rows(min_row=1, max_row=1))]
    prefix = f"{year}{month:02d}"
    user_rows = [
        list(row)
        for row in ws_src.iter_rows(min_row=2, values_only=True)
        if row[1] and user in str(row[1])
        and row[0] and str(int(row[0]) if isinstance(row[0], float) else row[0]).startswith(prefix)
    ]
    wb_src.close()
    if tmp and tmp.exists():
        tmp.unlink()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = XLSX_SHEET
    ws.append(header)
    for row in user_rows:
        ws.append(row)

    buf = tempfile.mktemp(suffix=".xlsx")
    wb.save(buf)
    data = Path(buf).read_bytes()
    Path(buf).unlink()
    return data


def send_summary_email(cfg: dict, rows: list[list], date_label: str, user: str,
                       shared_xlsx: Path | None = None) -> None:
    """Send a summary email with the monthly Excel attached."""
    # Load email config from dedicated file; fall back to main config for backwards compat
    email_cfg: dict = {}
    if EMAIL_CONFIG_FILE.exists():
        with open(EMAIL_CONFIG_FILE, "r", encoding="utf-8") as f:
            email_cfg = json.load(f)
    else:
        email_cfg = cfg  # backwards compat

    email_to = email_cfg.get("email_to", "").strip()
    smtp_host = email_cfg.get("smtp_host", "").strip()
    smtp_user = email_cfg.get("smtp_user", "").strip()
    smtp_password = email_cfg.get("smtp_password", "").strip()
    smtp_port = int(email_cfg.get("smtp_port", 587))

    if not email_to or not smtp_host or not smtp_user or not smtp_password:
        return  # email not configured — skip silently

    # Load ALL rows for the current month from the shared file (not just today's writes)
    today = date.today()
    month_rows = rows  # fallback to passed rows
    if shared_xlsx and shared_xlsx.exists():
        try:
            wb_r, tmp_r = _openpyxl_load(shared_xlsx, read_only=True)
            ws_r = wb_r[XLSX_SHEET]
            prefix = today.strftime("%Y%m")
            month_rows = [
                list(r) for r in ws_r.iter_rows(min_row=2, values_only=True)
                if r[1] and user in str(r[1])
                and r[0] and str(int(r[0]) if isinstance(r[0], float) else r[0]).startswith(prefix)
            ]
            wb_r.close()
            if tmp_r and tmp_r.exists():
                tmp_r.unlink()
        except Exception:
            month_rows = rows  # fallback

    # Build plain-text body
    lines = [
        f"Registo de horas — {date_label}",
        f"Utilizador: {user}",
        f"Mês: {today.strftime('%B %Y')}",
        "",
        f"{'Data':<12} {'Projeto':<15} {'DevOps':<8} {'Tarefa':<25} {'Horas':<6} Comentário",
        "-" * 90,
    ]
    total = 0.0
    for row in month_rows:
        date_col = row[8] if len(row) > 8 and row[8] else row[0]
        project_col = str(row[2])[:14]
        devops_col = str(row[3])[:7]
        task_col = str(row[4])[:24]
        try:
            h = float(row[5])
        except (ValueError, TypeError):
            h = 0.0
        if str(row[5]) != "N/A":
            total += h
        hours_col = str(row[5])
        comment_col = str(row[6])[:60]
        lines.append(f"{date_col:<12} {project_col:<15} {devops_col:<8} {task_col:<25} {hours_col:<6} {comment_col}")
    lines += [
        "-" * 90,
        f"Total: {total:.2f}h",
        "",
        "-- Enviado automaticamente pelo timesheet --",
    ]
    body = "\n".join(lines)

    today = date.today()
    month_name = today.strftime("%Y-%m")
    attachment_name = f"Timesheet_{user.replace(' ', '_')}_{month_name}.xlsx"

    msg = MIMEMultipart()
    msg["Subject"] = f"[Timesheet] Registo de horas {date_label} — {user}"
    msg["From"] = smtp_user
    msg["To"] = email_to
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Attach monthly Excel if shared file is available
    if shared_xlsx and shared_xlsx.exists():
        try:
            xlsx_bytes = _build_monthly_excel(shared_xlsx, user, today.year, today.month)
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(xlsx_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=attachment_name)
            msg.attach(part)
        except Exception as exc:
            print(f"[timesheet] AVISO: não foi possível gerar o anexo Excel — {exc}")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, [email_to], msg.as_bytes())
        print(f"[timesheet] Email de resumo enviado para {email_to} (anexo: {attachment_name}).")
    except Exception as exc:
        print(f"[timesheet] AVISO: não foi possível enviar o email — {exc}")


# ---------------------------------------------------------------------------
# Backfill helpers
# ---------------------------------------------------------------------------
def get_last_registered_date(xlsx_path: Path, user: str) -> "date | None":
    """Return the date of the last row registered for the given user in the xlsx."""
    if not xlsx_path.exists():
        return None
    wb, tmp = _openpyxl_load(xlsx_path, read_only=True)
    ws = wb[XLSX_SHEET]
    result = _ws_get_last_registered_date(ws, user)
    wb.close()
    if tmp and tmp.exists():
        tmp.unlink()
    return result


def get_dates_to_register(last_date: "date | None", target_date: date) -> list:
    """
    Return all weekdays (Mon-Fri) from last_date+1 up to target_date inclusive.
    If last_date is None, returns only [target_date].
    """
    if last_date is None:
        return [target_date]
    dates = []
    current = last_date + timedelta(days=1)
    while current <= target_date:
        if current.weekday() < 5:  # Mon-Fri only
            dates.append(current)
        current += timedelta(days=1)
    return dates if dates else [target_date]


# ---------------------------------------------------------------------------
# Windows Task Scheduler
# ---------------------------------------------------------------------------
TASK_NAME = "TimesheetDailyRegisto"


def schedule_task(time_str: str) -> None:
    """Create or replace a Windows Task Scheduler task to run this script Mon-Fri."""
    import re as _re
    if not _re.match(r"^\d{2}:\d{2}$", time_str):
        print(f"[timesheet] Erro: formato de hora inválido '{time_str}'. Usa HH:MM, ex: 18:00.")
        sys.exit(1)

    python_exe = sys.executable
    script_path = str(Path(__file__).resolve())

    # schtasks days: MON,TUE,WED,THU,FRI
    cmd = [
        "schtasks", "/Create", "/F",
        "/TN", TASK_NAME,
        "/TR", f'"{python_exe}" "{script_path}"',
        "/SC", "WEEKLY",
        "/D", "MON,TUE,WED,THU,FRI",
        "/ST", time_str,
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode == 0:
        print(f"[timesheet] Tarefa '{TASK_NAME}' agendada para {time_str} (seg-sex).")
        print(f"            Executável: {python_exe}")
        print(f"            Script:     {script_path}")
    else:
        print(f"[timesheet] Erro ao criar tarefa agendada:\n{result.stderr or result.stdout}")
        sys.exit(1)


def unschedule_task() -> None:
    """Remove the scheduled task from Windows Task Scheduler."""
    cmd = ["schtasks", "/Delete", "/F", "/TN", TASK_NAME]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode == 0:
        print(f"[timesheet] Tarefa '{TASK_NAME}' removida.")
    else:
        print(f"[timesheet] Erro ao remover tarefa (já não existe?):\n{result.stderr or result.stdout}")


# ---------------------------------------------------------------------------
# Local copy sync
# ---------------------------------------------------------------------------
def _sync_local_copy(shared_xlsx: Path, local_xlsx: Path, user: str) -> None:
    """
    Rebuild the local Excel file with all rows for `user` from the shared file,
    using the same column structure. Called after every successful write.
    """
    # Read header + user rows from shared file
    wb_src, tmp = _openpyxl_load(shared_xlsx, read_only=True)
    ws_src = wb_src[XLSX_SHEET]

    header = [cell.value for cell in next(ws_src.iter_rows(min_row=1, max_row=1))]

    user_rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[1] and user in str(row[1]):
            user_rows.append(list(row))

    wb_src.close()
    if tmp and tmp.exists():
        tmp.unlink()

    # Write local file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = XLSX_SHEET
    ws.append(header)
    for row in user_rows:
        ws.append(row)
    wb.save(local_xlsx)
    print(f"[timesheet] Cópia local actualizada: {local_xlsx.name} ({len(user_rows)} registos)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Regista horas diárias no timesheet a partir de commits git."
    )
    parser.add_argument(
        "--date",
        metavar="YYYYMMDD",
        help="Data a registar (por omissão: hoje)",
    )
    parser.add_argument(
        "--config",
        action="store_true",
        help="Reconfigurar utilizador e definições",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Mostrar o que seria escrito sem escrever no CSV",
    )
    parser.add_argument(
        "--schedule",
        metavar="HH:MM",
        help="Agendar execução diária (seg-sex) a esta hora, ex: 18:00. "
             "Cria/actualiza uma tarefa no Windows Task Scheduler.",
    )
    parser.add_argument(
        "--unschedule",
        action="store_true",
        help="Remover a tarefa agendada do Windows Task Scheduler.",
    )
    args = parser.parse_args()

    # Load / prompt config
    cfg = ensure_config(force=args.config)
    if args.config:
        print("[timesheet] Configuração atualizada. Corre sem --config para registar horas.")
        return

    if args.unschedule:
        unschedule_task()
        return

    if args.schedule:
        schedule_task(args.schedule)
        return

    user = cfg["user"]
    project = cfg.get("project", "SI Pessoas")
    task = cfg.get("default_task", "Dev")
    total_hours = float(cfg.get("default_hours", 8.0))
    # Support both old single-repo and new multi-repo config
    git_repos_raw = cfg.get("git_repos", cfg.get("git_repo_path", "."))
    git_repos: list[str] = git_repos_raw if isinstance(git_repos_raw, list) else [git_repos_raw]
    git_author = cfg.get("git_author", "")
    xlsx_file = Path(cfg["xlsx_file"]) if cfg.get("xlsx_file") else XLSX_FILE

    # Resolve list of dates to register
    if args.date:
        try:
            explicit_date = datetime.strptime(args.date, "%Y%m%d").date()
        except ValueError:
            print(f"[timesheet] Erro: formato de data inválido '{args.date}'. Usa YYYYMMDD.")
            sys.exit(1)
        if explicit_date.weekday() >= 5:
            weekday_name = WEEKDAY_PT[explicit_date.weekday()]
            print(f"[timesheet] {explicit_date.strftime('%d/%m/%Y')} é {weekday_name} — sem registo ao fim de semana.")
            return
        dates_to_register = [explicit_date]
    else:
        today = date.today()
        if today.weekday() >= 5:
            weekday_name = WEEKDAY_PT[today.weekday()]
            print(f"[timesheet] {today.strftime('%d/%m/%Y')} é {weekday_name} — sem registo ao fim de semana.")
            return
        last_date = get_last_registered_date(xlsx_file, user)
        dates_to_register = get_dates_to_register(last_date, today)
        if last_date:
            print(f"[timesheet] Último registo: {last_date.strftime('%d/%m/%Y')}. "
                  f"A verificar {len(dates_to_register)} dia(s) em falta...")
        else:
            print(f"[timesheet] Sem registos anteriores. A registar hoje ({today.strftime('%d/%m/%Y')}).")

    if not dates_to_register:
        print("[timesheet] Nenhum dia em falta.")
        return

    all_written_rows: list = []

    if args.dry_run:
        for target_date in dates_to_register:
            date_label = target_date.strftime("%d/%m/%Y")
            commits = get_commits_for_date_extended(git_repos, target_date, git_author)
            commit_summary = _commit_summary(commits)
            print(f"[timesheet] {date_label} — {len(commits)} commit(s)" +
                  (f": {commit_summary}" if commit_summary else " → N/A"))
            rows = build_rows(target_date, user, project, task, total_hours, commits)
            for row in rows:
                print("  [dry-run] " + " | ".join(str(c) for c in row))
        return

    # --- Single openpyxl session for all dates ---
    wb, _tmp_src = _openpyxl_load(xlsx_file)
    ws = wb[XLSX_SHEET]
    existing = _ws_read_existing_dates(ws)

    for target_date in dates_to_register:
        date_label = target_date.strftime("%d/%m/%Y")
        date_str = target_date.strftime("%Y%m%d")
        commits = get_commits_for_date_extended(git_repos, target_date, git_author)
        commit_summary = _commit_summary(commits)
        print(f"[timesheet] {date_label} — {len(commits)} commit(s)" +
              (f": {commit_summary}" if commit_summary else " → N/A"))
        rows = build_rows(target_date, user, project, task, total_hours, commits)
        duplicate_key = f"{date_str}|{user}"
        if duplicate_key in existing:
            deleted = _ws_delete_rows_for_date(ws, date_str, user)
            _ws_append_rows(ws, rows)
            print(f"  [actualizado] substituiu {deleted} linha(s) existente(s)")
        else:
            _ws_append_rows(ws, rows)
        for row in rows:
            print(f"  + {row[3]:<8} {row[5]}h  {row[6][:55]}")
        all_written_rows.extend(rows)
        existing.add(duplicate_key)  # prevent double-write if date appears again

    _openpyxl_save(wb, xlsx_file, _tmp_src)
    _sync_local_copy(xlsx_file, XLSX_LOCAL, user)

    total_written = len(all_written_rows)
    if total_written:
        print(f"\n[timesheet] {total_written} linha(s) adicionada(s) no total.")
        first_label = dates_to_register[0].strftime("%d/%m/%Y")
        last_label = dates_to_register[-1].strftime("%d/%m/%Y")
        email_label = last_label if len(dates_to_register) == 1 else f"{first_label} a {last_label}"
        send_summary_email(cfg, all_written_rows, email_label, user, shared_xlsx=xlsx_file)
    else:
        print("[timesheet] Nenhuma linha nova registada.")
        send_summary_email(cfg, [], date.today().strftime("%d/%m/%Y"), user, shared_xlsx=xlsx_file)


if __name__ == "__main__":
    main()
